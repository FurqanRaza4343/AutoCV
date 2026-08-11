import json
import os
import io
import csv
import threading
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..database import SessionLocal, get_db
from .. import models, schemas
from ..pipeline_agents import run_pipeline_parallel
from .auth import get_current_user

router = APIRouter(prefix="/pipeline", tags=["pipeline"])


def _background_run_pipeline(run_id: str, candidate_ids: list[str], job_title: str, job_description: str):
    # Fetch candidates and close this session immediately - it must not sit open and
    # idle for the whole parse+screen duration (which can run tens of seconds across
    # several Mistral calls). A connection held idle-but-checked-out for that long can
    # go stale server-side (managed Postgres proxy idle timeout) in a way pool_pre_ping
    # doesn't catch (it only re-validates connections when they're re-acquired FROM the
    # pool, not ones a session is still actively holding) - the previous version reused
    # this same long-idle connection for the post-scoring commit, which is exactly where
    # pipeline runs were hanging indefinitely.
    db = SessionLocal()
    try:
        run = db.query(models.PipelineRun).filter(models.PipelineRun.id == run_id).first()
        if not run:
            return

        candidates_map = {}
        for cid in candidate_ids:
            c = db.query(models.Candidate).filter(models.Candidate.id == cid).first()
            if c:
                candidates_map[cid] = c
    finally:
        db.close()

    total = len(candidate_ids)

    def update_progress(done: int, total: int):
        # Runs once per candidate from inside the parse+screen thread pool - every
        # session opened must be closed unconditionally, not just on the success path.
        udb = SessionLocal()
        try:
            r = udb.query(models.PipelineRun).filter(models.PipelineRun.id == run_id).first()
            if r:
                r.progress = min(10 + int(done / total * 55), 65)
                r.parsed_count = done
                r.screened_count = done
                r.current_agent = "Parser + Screener Agent"
                udb.commit()
        except Exception:
            pass
        finally:
            udb.close()

    working_data, best_wd = run_pipeline_parallel(candidates_map, candidate_ids, job_description, progress_callback=update_progress)

    # Fresh session for the write-back phase - the one above is long gone by now.
    db = SessionLocal()
    try:
        run = db.query(models.PipelineRun).filter(models.PipelineRun.id == run_id).first()
        if not run:
            return

        run.current_agent = "Deep Ranker + Finalizer Agent"
        run.progress = 70
        db.commit()

        db.query(models.PipelineResult).filter(models.PipelineResult.run_id == run.id).delete()
        db.commit()

        # best_wd (from agent_finalize via run_pipeline_parallel) is only ever a
        # candidate that was actually scored (status == "ok") - it's None when every
        # result in this batch is a lead with no CV or a failed AI call, in which case
        # nothing should be crowned "Best Match" (a bare `working_data[0]` pick here
        # previously let an unscored lead show up as the best match with a blank score).
        best_candidate_id = best_wd.get("candidate_id") if best_wd else None

        for idx, wd in enumerate(working_data):
            p = wd.get("parsed", {})
            s = wd.get("screened", {})
            r = wd.get("ranked", {})
            f = wd.get("final", {})
            skills_list = p.get("skills", [])
            skills_str = ", ".join(skills_list) if isinstance(skills_list, list) else str(skills_list)

            is_best = best_candidate_id is not None and wd["candidate_id"] == best_candidate_id

            # agent_finalize() also computes a concrete "what to do about this" next step
            # (e.g. "ask for a resume and re-run" for a lead) - there's no dedicated column
            # for it, so fold it into final_notes rather than silently dropping it; this is
            # the one field surfaced everywhere a result is shown (UI card + all 3 exports).
            final_notes_text = f.get("final_notes", "")
            next_steps = f.get("next_steps", "")
            if next_steps:
                final_notes_text = f"{final_notes_text} Next step: {next_steps}".strip()

            result = models.PipelineResult(
                run_id=run.id,
                candidate_id=wd["candidate_id"],
                candidate_name=p.get("name", wd["candidate"].get("name", "Unknown")),
                candidate_email=p.get("email", wd["candidate"].get("email", "")),
                role=p.get("role", wd["candidate"].get("role", "")),
                parsed_skills=skills_str,
                parsed_experience=p.get("experience_years"),
                parsed_location=p.get("location", ""),
                screened_score=s.get("screened_score"),
                screened_summary=s.get("screened_summary", ""),
                
                ranked_score=r.get("ranked_score"),
                ranked_analysis=r.get("ranked_analysis", ""),
                rank_position=r.get("rank_position", idx + 1),
                final_verdict=f.get("final_verdict", "Consider"),
                final_notes=final_notes_text,
                is_best_match=is_best,
            )
            db.add(result)

            # candidates_map holds objects loaded in the earlier (now-closed) session -
            # they're detached, so mutating them wouldn't be tracked by this session.
            # Re-fetch by id instead of reusing the stale reference.
            orig = db.query(models.Candidate).filter(models.Candidate.id == wd["candidate_id"]).first()
            if orig:
                orig.match_score = r.get("ranked_score") or s.get("screened_score") or orig.match_score
                orig.current_stage = "Done"
                orig.summary = s.get("screened_summary", orig.summary)

        run.status = "completed"
        run.progress = 100
        run.current_agent = None
        run.completed_at = datetime.now(timezone.utc)
        db.commit()

    except Exception as e:
        try:
            run = db.query(models.PipelineRun).filter(models.PipelineRun.id == run_id).first()
            if run:
                run.status = "failed"
                run.error_message = str(e)[:500]
                db.commit()
        except Exception:
            pass
    finally:
        db.close()


@router.post("/run")
def run_pipeline(payload: schemas.PipelineRunCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    job_title = payload.job_title or "Software Engineer"
    job_description = payload.job_description or ""
    candidate_ids = payload.candidate_ids

    if not candidate_ids:
        all_candidates = db.query(models.Candidate).filter(models.Candidate.user_id == current_user.id).order_by(models.Candidate.created_at.desc()).limit(50).all()
        candidate_ids = [c.id for c in all_candidates]
    else:
        # Explicit ids from the client are still constrained to this user's own
        # candidates - never let a request run the pipeline over someone else's data.
        owned = db.query(models.Candidate.id).filter(models.Candidate.id.in_(candidate_ids), models.Candidate.user_id == current_user.id).all()
        candidate_ids = [c.id for c in owned]

    if not candidate_ids:
        raise HTTPException(400, "No candidates found to run the pipeline. Fetch or upload candidates first.")

    run = models.PipelineRun(
        user_id=current_user.id,
        job_title=job_title,
        job_description=job_description,
        status="running",
        total_candidates=len(candidate_ids),
    )
    db.add(run)
    db.commit()
    db.refresh(run)

    threading.Thread(
        target=_background_run_pipeline,
        args=(run.id, candidate_ids, job_title, job_description),
        daemon=True,
    ).start()

    return {
        "run": schemas.PipelineRunOut.model_validate(run),
        "results": [],
        "message": "Pipeline started in background. Poll GET /api/pipeline/runs/{id} for status.",
    }


@router.get("/runs", response_model=list[schemas.PipelineRunOut])
def list_pipeline_runs(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    return db.query(models.PipelineRun).filter(models.PipelineRun.user_id == current_user.id).order_by(models.PipelineRun.created_at.desc()).all()


@router.get("/runs/{run_id}", response_model=schemas.PipelineFullOut)
def get_pipeline_run(run_id: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    run = db.query(models.PipelineRun).filter(models.PipelineRun.id == run_id, models.PipelineRun.user_id == current_user.id).first()
    if not run:
        raise HTTPException(404, "Pipeline run not found")
    results = db.query(models.PipelineResult).filter(models.PipelineResult.run_id == run_id).order_by(models.PipelineResult.rank_position).all()
    return schemas.PipelineFullOut(
        run=schemas.PipelineRunOut.model_validate(run),
        results=[schemas.PipelineResultOut.model_validate(r) for r in results],
    )


@router.get("/runs/{run_id}/results")
def get_pipeline_results(run_id: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    run = db.query(models.PipelineRun).filter(models.PipelineRun.id == run_id, models.PipelineRun.user_id == current_user.id).first()
    if not run:
        raise HTTPException(404, "Pipeline run not found")
    results = db.query(models.PipelineResult).filter(models.PipelineResult.run_id == run_id).order_by(models.PipelineResult.rank_position).all()
    return [schemas.PipelineResultOut.model_validate(r) for r in results]


def _get_owned_result(db: Session, result_id: str, user_id: str) -> models.PipelineResult:
    """PipelineResult has no user_id of its own - ownership is via its PipelineRun."""
    r = (
        db.query(models.PipelineResult)
        .join(models.PipelineRun, models.PipelineResult.run_id == models.PipelineRun.id)
        .filter(models.PipelineResult.id == result_id, models.PipelineRun.user_id == user_id)
        .first()
    )
    if not r:
        raise HTTPException(404, "Result not found")
    return r


@router.get("/results/{result_id}/export-txt")
def export_result_txt(result_id: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    r = _get_owned_result(db, result_id, current_user.id)

    lines = [
        "=" * 60,
        f"AGENTIX AI - CANDIDATE EVALUATION REPORT",
        "=" * 60,
        "",
        f"Candidate: {r.candidate_name}",
        f"Email: {r.candidate_email or 'N/A'}",
        f"Role: {r.role or 'N/A'}",
        "",
        "-" * 40,
        "PARSED DATA",
        "-" * 40,
        f"Skills: {r.parsed_skills or 'N/A'}",
        f"Experience: {r.parsed_experience or 'N/A'} years",
        f"Location: {r.parsed_location or 'N/A'}",
        "",
        "-" * 40,
        "SCREENING RESULTS",
        "-" * 40,
        f"Score: {r.screened_score or 'N/A'}/100",
        f"Summary: {r.screened_summary or 'N/A'}",
        "",
        "-" * 40,
        "RANKING RESULTS",
        "-" * 40,
        f"Rank Position: #{r.rank_position or 'N/A'}",
        f"Ranked Score: {r.ranked_score or 'N/A'}/100",
        f"Analysis: {r.ranked_analysis or 'N/A'}",
        "",
        "-" * 40,
        "FINAL DECISION",
        "-" * 40,
        f"Verdict: {r.final_verdict or 'N/A'}",
        f"Notes: {r.final_notes or 'N/A'}",
        f"Best Match: {'Yes' if r.is_best_match else 'No'}",
        "",
        "=" * 60,
        "Generated by Agentix AI HR Manager",
        "=" * 60,
    ]
    content = "\n".join(lines)
    return StreamingResponse(
        io.StringIO(content),
        media_type="text/plain",
        headers={"Content-Disposition": f"attachment; filename={r.candidate_name.replace(' ', '_')}_report.txt"},
    )


@router.get("/results/{result_id}/export-xlsx")
def export_result_xlsx(result_id: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    r = _get_owned_result(db, result_id, current_user.id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Candidate Report"

    ws.append(["AGENTIX AI - CANDIDATE EVALUATION REPORT"])
    ws.append([])
    ws.append(["Field", "Value"])
    ws.append(["Candidate", r.candidate_name])
    ws.append(["Email", r.candidate_email or "N/A"])
    ws.append(["Role", r.role or "N/A"])
    ws.append(["Skills", r.parsed_skills or "N/A"])
    ws.append(["Experience (years)", r.parsed_experience or "N/A"])
    ws.append(["Location", r.parsed_location or "N/A"])
    ws.append(["Screened Score", r.screened_score or "N/A"])
    ws.append(["Screening Summary", r.screened_summary or "N/A"])
    ws.append(["Rank Position", r.rank_position or "N/A"])
    ws.append(["Ranked Score", r.ranked_score or "N/A"])
    ws.append(["Rank Analysis", r.ranked_analysis or "N/A"])
    ws.append(["Final Verdict", r.final_verdict or "N/A"])
    ws.append(["Final Notes", r.final_notes or "N/A"])
    ws.append(["Best Match", "Yes" if r.is_best_match else "No"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={r.candidate_name.replace(' ', '_')}_report.xlsx"},
    )


@router.get("/results/{result_id}/export-pdf")
def export_result_pdf(result_id: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    try:
        from fpdf import FPDF
    except ImportError:
        try:
            from fpdf2 import FPDF
        except ImportError:
            raise HTTPException(500, "fpdf2 not installed. Run: pip install fpdf2")

    r = _get_owned_result(db, result_id, current_user.id)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Agentix AI - Candidate Evaluation Report", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(10)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, f"Candidate: {r.candidate_name}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Email: {r.candidate_email or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"Role: {r.role or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "Parsed Data", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.multi_cell(0, 5, f"Skills: {r.parsed_skills or 'N/A'}")
    pdf.cell(0, 5, f"Experience: {r.parsed_experience or 'N/A'} years", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, f"Location: {r.parsed_location or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "Screening Results", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, f"Score: {r.screened_score or 'N/A'}/100", new_x="LMARGIN", new_y="NEXT")
    pdf.multi_cell(0, 5, f"Summary: {r.screened_summary or 'N/A'}")
    pdf.ln(3)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "Ranking Results", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, f"Rank Position: #{r.rank_position or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, f"Ranked Score: {r.ranked_score or 'N/A'}/100", new_x="LMARGIN", new_y="NEXT")
    pdf.multi_cell(0, 5, f"Analysis: {r.ranked_analysis or 'N/A'}")
    pdf.ln(3)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "Final Decision", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, f"Verdict: {r.final_verdict or 'N/A'}", new_x="LMARGIN", new_y="NEXT")
    pdf.multi_cell(0, 5, f"Notes: {r.final_notes or 'N/A'}")
    pdf.cell(0, 5, f"Best Match: {'Yes' if r.is_best_match else 'No'}", new_x="LMARGIN", new_y="NEXT")

    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={r.candidate_name.replace(' ', '_')}_report.pdf"},
    )
